from flask import render_template, request, redirect, url_for
from app import app
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from apscheduler.schedulers.background import BackgroundScheduler
from plyer import notification
import os

# In-memory storage for patient details
appointments = []
records = []

# Path to the Excel file
EXCEL_FILE = "patient_records.xlsx"

# Function to initialize the Excel file
def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Patient Records"
        # Adding headers
        sheet.append(["Patient Name", "Contact", "Appointment Time", "Address", "Appointment For", "Status"])
        workbook.save(EXCEL_FILE)

# Function to load records from Excel at startup
def load_records_from_excel():
    if os.path.exists(EXCEL_FILE):
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):  
             if len(row) == 6:  # Skip the header row
                 patient_details = {
                    "patient_name": row[0],
                    "patient_contact": row[1],
                    "appointment_time": datetime.strptime(row[2], "%Y-%m-%d %H:%M"),
                    "address": row[3],
                    "appointment_for": row[4],
                    "status": row[5]
                }
                 appointments.append(patient_details)
                 records.append(patient_details)

# Initialize the Excel file and load records
initialize_excel()
load_records_from_excel()

# Function to send system notifications
def send_notification(patient_name, appointment_time):
    notification.notify(
        title="Appointment Reminder",
        message=f"Reminder: {patient_name} has an appointment at {appointment_time}.",
        timeout=10  # Notification duration in seconds
    )

# Function to check appointments and send reminders
def check_appointments():
    now = datetime.now()
    for appointment in appointments:
        appointment_time = appointment['appointment_time']
        if now + timedelta(hours=2) >= appointment_time > now:
            send_notification(appointment['patient_name'], appointment_time.strftime("%Y-%m-%d %H:%M"))

# Setup the APScheduler
scheduler = BackgroundScheduler()
scheduler.add_job(check_appointments, 'interval', minutes=1)  # Run every minute
scheduler.start()

@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        # Get form data
        patient_name = request.form.get("patient_name")
        patient_contact = request.form.get("patient_contact")
        appointment_time = request.form.get("appointment_time")
        address = request.form.get("address")
        appointment_for = request.form.get("appointment_for")
        status = request.form.get("status")  # Get status field

        # Convert the appointment time to a datetime object
        appointment_time = datetime.strptime(appointment_time, "%Y-%m-%dT%H:%M")

        # Patient details dictionary
        patient_details = {
            "patient_name": patient_name,
            "patient_contact": patient_contact,
            "appointment_time": appointment_time,
            "address": address,
            "appointment_for": appointment_for,
            "status": status  # Add status to the record
        }

        # Append details to in-memory storage
        appointments.append(patient_details)
        records.append(patient_details)

        # Write to Excel file
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
        sheet.append([
            patient_name,
            patient_contact,
            appointment_time.strftime("%Y-%m-%d %H:%M"),
            address,
            appointment_for,
            status  # Add status to the Excel record
        ])
        workbook.save(EXCEL_FILE)

        return redirect(url_for("appointments_view"))

    return render_template("home.html")


@app.route("/appointments")
def appointments_view():
      # Sort appointments by appointment time
    sorted_appointments = sorted(appointments, key=lambda x: x['appointment_time'])
    # Group appointments by day
    appointments_by_day = {}
    for appointment in appointments:
        day = appointment['appointment_time'].date()
        if day not in appointments_by_day:
            appointments_by_day[day] = []
        appointments_by_day[day].append(appointment)
    
      # Sort the days to display today's appointments first
    sorted_days = sorted(appointments_by_day.keys())
    return render_template("appointments.html", appointments_by_day=appointments_by_day,sorted_days=sorted_days)


@app.route("/records", methods=["GET"])
def records_view():
    search_query = request.args.get('search', '').lower()  # Get search query, default to empty string
    filtered_records = []

    # If there's a search query, filter the records
    if search_query:
        for record in records:
            if search_query in record['patient_name'].lower() or search_query in record['patient_contact'].lower():
                filtered_records.append(record)
    else:
        filtered_records = records  # Show all records if no search query

    return render_template("records.html", records=filtered_records)
